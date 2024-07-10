from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell
from copy import copy
import numpy as np


"""
How to modify the exporter : 
- 1. Adjust `VERT_SPACING`and `HORI_SPACING` according to you needs
- 2. Insert the correct data into the company_regional_data dictionary located in generate_data()
- 3. Modify data_to_matrix() to insert the data at the desired place
- 4. Modify the labels in generate_workbook(). These are the labels on the left of the output sheet.
- 5. Modify add_borders_to_unit() to add borders around the newly inserted spots. Don't forget to modify the other thin borders if the one you insert isn't at the bottom. Ex = `add_thin_border(ws[f'{get_column_letter(start_col+1)}{start_row+16}'])` where you would modify 16 by the number of added rows (eg. 16+3)
"""


VERT_SPACING = 19
HORI_SPACING = 3

def generate_data(session) -> dict:
    assert session.n_regions == 1, NotImplementedError(f'Multiple regions are still unsupported, but n_regions = {session.n_regions} was given')
    data = {}
    data['n_companies'] = len(session.marketPlayers)
    data['n_regions'] = session.n_regions
    data['Quarter'] = session.quarter
    for company in session.marketPlayers:
        label = f'Company{company.id}'
        data[label] = {}
        data[label]['Region1'] = {}
        company_regional_data = data[label]['Region1']
        company_regional_data['Inventory X'] =      company.get_inventory('X')
        company_regional_data['Inventory Y'] =      company.get_inventory('Y')
        company_regional_data['Sales_X'] =          company.get_inventory('X', type_='sales')
        company_regional_data['Sales_Y'] =          company.get_inventory('Y', type_='sales')
        company_regional_data['Production_X'] =     company.get_inventory('X', type_='production')
        company_regional_data['Production_Y'] =     company.get_inventory('Y', type_='production')
        company_regional_data['B2B_Sales_X'] = aggregate_B2B_sales_by_grade(session.transactions, company, session.quarter, 'X')
        company_regional_data['B2B_Sales_Y'] = aggregate_B2B_sales_by_grade(session.transactions, company, session.quarter, 'Y')

        ##
        company_regional_data['Max Grade X'] = company.max_grades['X']
        company_regional_data['Max Grade Y'] = company.max_grades['Y']
        
        # Getting factory ages and number
        for item in ('X', 'Y'):
            N_Factories = 0
            ages = '['
            for _, factory in company.factories[item].items():
                ages += str(int(factory.age))
                ages += ', '
                N_Factories += 1
            ages = ages[:-2] + ']' if ages != '[' else 'Empty' # Removes the extra ', '
            company_regional_data[f'Factory {item} age'] = ages
            company_regional_data[f'N° Factories {item}'] = N_Factories
        company_regional_data['N° Sales Office'] = company.n_sales_offices
    return data

def data_to_matrix(period_data:dict) -> np.ndarray:
    p = period_data
    fmt_X_main = format_inventory(p['Inventory X'])
    fmt_Y_main = format_inventory(p['Inventory Y'])

    fmt_X_sales = format_inventory(p['Sales_X'])
    fmt_Y_sales = format_inventory(p['Sales_Y'])

    fmt_X_prod = format_inventory(p['Production_X'])
    fmt_Y_prod = format_inventory(p['Production_Y'])


    matrix = [
        [fmt_X_main[0],fmt_X_main[1],None],
        [fmt_Y_main[0],fmt_Y_main[1],None],
        [None,None,None],
        [fmt_X_sales[0],fmt_X_sales[1],None],
        [fmt_Y_sales[0],fmt_Y_sales[1],None],
        [None,None,None],
        [p['B2B_Sales_X'][0],p['B2B_Sales_X'][1],None],
        [p['B2B_Sales_Y'][0],p['B2B_Sales_Y'][1],None],
        [None,None,None],
        [fmt_X_prod[0],fmt_X_prod[1], None],
        [fmt_Y_prod[0],fmt_Y_prod[1], None],
        [None,None,None],
        [p['Max Grade X'],None, None],
        [p['Max Grade Y'], None, None],
        [None,None,None],
        [p['N° Factories X'], p['Factory X age'], None],
        [p['N° Factories Y'], p['Factory Y age'], None],
        [None, None, None],
        [p['N° Sales Office']]
    ]
    return matrix

def generate_workbook(data) -> Workbook:
    n_regions = data['n_regions']
    n_companies = data['n_companies']
    if data["Quarter"] != 1 :
        wb = load_workbook('output.xlsx')
    else :
        wb = Workbook()
    wb.active = wb.create_sheet(title=f"Period{data['Quarter']}")
    ws = wb.active

    col_start = 3
    # Setting up the company names in worksheet
    for cid in range(1, n_companies + 1):
        col_start_letter = get_column_letter(col_start)
        col_end_letter = get_column_letter(col_start + HORI_SPACING -1)
        ws.merge_cells(f'{col_start_letter}{1}:{col_end_letter}{2}')
        new_cell_ref = f"{col_start_letter}{1}"
        ws[new_cell_ref].value = f'Company {cid}'
        ws[new_cell_ref].font = Font(bold=True, size=24)

        col_start += HORI_SPACING + 1

    ws['B4'] = 'Company ID'
    labels = ['Inventory X (u)', 'Inventory Y (u)', '',
              'Sales_X (u)', 'Sales_Y (u)', '', 
              'B2B_Sales_X (u)', 'B2B_Sales_Y (u)', '', 
              'Production_X (u)', 'Production_Y (u)', '', 
              'Max_Grade_X', 'Max_Grade_Y','', 
              'Factory X | Age', 'Factory Y | Age', ''
              , 'N° Sales Offices']
    
    # Merge cells in Column A and set "Region" value
    start_row = 6
    for i in range(n_regions):
        end_row = start_row + VERT_SPACING

        # Creating the Region Label
        merge_range = f"A{start_row}:A{end_row - 1}"
        ws.merge_cells(merge_range)
        ws[f'A{start_row}'].value = f'Region {i}'
        ws[f'A{start_row}'].font = Font(bold=True, size=24)
        ws[f'A{start_row}'].alignment = Alignment(textRotation=90, horizontal='center', vertical='center')
        
        # Populating the label columns
        for row, value in zip(range(start_row, end_row), labels):
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = Font(size=12)  # Optional: Set font size for label columns

        # Apply thick border to labels column
        border_sides = Border(left=Side(border_style='thick'), right=Side(border_style='thick'))
        border_top = Border(left=Side(border_style='thick'), right=Side(border_style='thick'), top=Side(border_style='thick'))
        border_bottom = Border(left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        
        for row in ws[f'B{start_row}:B{end_row-1}']:
            for cell in row:
                cell.border = border_sides

        ws[f'B{start_row}'].border = border_top
        ws[f'B{end_row -1}'].border = border_bottom

        start_row = end_row + 1

        # Adjusting column size
        ws.column_dimensions['B'].width = 15

    return wb

def populate_workbook(wb:Workbook, data) -> Workbook:
    n_companies = data["n_companies"]
    n_regions = data["n_regions"]
    

    # Populate company data
    start_column = 3
    for cid in range(1,n_companies+1):
        # Populate Unit
        company_data = data[f'Company{cid}']
        start_row = 6
        ws = wb.active
        ws.cell(row=4, column=start_column, value=cid)
        for rid in range(1, n_regions+1):
            region_data_matrix = data_to_matrix(company_data[f'Region{rid}'])
            wb = populate_unit(wb=wb, matrix = region_data_matrix, start_column=start_column, start_row=start_row)
            start_row += VERT_SPACING +1
        start_column += HORI_SPACING +1
    return wb

def populate_header(wb:Workbook, cid:int, start_column:int, start_row:int):
    ws = wb.active
    start_column_letter = get_column_letter(start_column)
    ws[f'{start_column_letter}{start_row}'] = cid
    return wb

def populate_unit(wb:Workbook,matrix, start_column, start_row):
    ws = wb.active

    for row_index, row in enumerate(matrix):
        for col_index, value in enumerate(row):
            cell = ws.cell(row=start_row + row_index, column=start_column + col_index)
            cell.value = value

    # wb = apply_border_to_square(wb, start_column, start_row, len(matrix[0]), len(matrix))
    return wb

def style_borders(wb:Workbook, data) -> Workbook:
    n_companies = data['n_companies']
    n_regions = data['n_regions']
    ws = wb.active

    col_start = 3

    # Placing thick borders around the companies names
    for _ in range(1, n_companies + 1):
        col_start_letter = get_column_letter(col_start)
        col_end_letter = get_column_letter(col_start + HORI_SPACING -1)
        merge_range = f'{col_start_letter}{1}:{col_end_letter}{2}'
        ws.merge_cells(merge_range)
            # Apply thick border to merged cells
        border = Border(
            left=Side(border_style="thick"),
            right=Side(border_style="thick"),
            top=Side(border_style="thick"),
            bottom=Side(border_style="thick")
        )
        for row in ws[merge_range]:
            for cell in row:
                cell.border = border

        col_start += HORI_SPACING + 1

    # Adding thin borders
    row_start = 6
    for _ in range(0,n_regions):
        col_start = 3
        for col_index in range(col_start, col_start + 4*n_companies, 4):
            apply_border_to_square(wb, col_index, row_start, HORI_SPACING, VERT_SPACING)
            add_thin_borders_to_unit(wb, col_index, row_start)
        row_start += VERT_SPACING + 2
        
    # Adding header borders
    col_start = 3
    for _ in range (0,4*n_companies, 4):
        
        top_left_border = Border(top=Side(border_style="thick"), left = Side(border_style="thick"))
        top_right_border = Border(top=Side(border_style="thick"), right = Side(border_style="thick"))
        bottom_left_border = Border(bottom=Side(border_style="thick"), left =Side(border_style="thick"))
        bottom_right_border = Border(bottom=Side(border_style="thick"), right=Side(border_style="thick"))
        left_border = Border(left=Side(border_style="thick"))
        right_border = Border(right=Side(border_style="thick"))
        
        ws[f'{get_column_letter(col_start)}{3}'].border = top_left_border
        ws[f'{get_column_letter(col_start)}{4}'].border = left_border
        ws[f'{get_column_letter(col_start)}{5}'].border = bottom_left_border

        ws[f'{get_column_letter(col_start+2)}{3}'].border = top_right_border
        ws[f'{get_column_letter(col_start+2)}{4}'].border = right_border
        ws[f'{get_column_letter(col_start+2)}{5}'].border = bottom_right_border

        add_thin_border(ws[f'{get_column_letter(col_start)}{4}'])

        col_start += 4

    # Disactivating gridlines view
    wb.active.sheet_view.showGridLines = False
    return wb

def export_data(session)-> None:
    """Exports the data from the current session and adds it to './output.xlsx

    Parameters
    ----------
    session: Session
        The session from which to export the data

    Returns
    -------
    None

    """
    period_data = generate_data(session)
    wb = generate_workbook(period_data)
    wb = populate_workbook(wb,period_data)
    wb = style_borders(wb, period_data)
    
    try:
        wb.remove(wb['Sheet'])
    except:
        pass
    wb.save('output.xlsx')

def format_inventory(inventory:np.ndarray) -> str:
    indices = inventory.nonzero()[0]      
    values = inventory[indices]
    formatted = []
    for index, value in list(zip(indices, values)) : 
        formatted.append(f'({index}) {int(value)}')
    if len(formatted) == 1 : 
        formatted.append('Empty')
    if len(formatted) == 0 :
        formatted = ['Empty', 'Empty']
    return formatted

###########
# Helpers #
###########

def aggregate_B2B_sales_by_grade(registry, company, quarter, item):
    """
    Calculate the aggregate volume of Business-to-Business (B2B) sales for a specific company, quarter, and item,
    grouped by grade.

    Parameters:
    - registry (pd.DataFrame): The DataFrame containing sales data.
    - company (int): The company ID (seller) for which to calculate sales.
    - quarter (int): The quarter for which to calculate sales.
    - item (str): The item or product type for which to calculate sales.

    Returns:
    - list: A list of strings, where each string represents a grade with its corresponding volume in the format '(grade) volume'.

    Example:
    >>> df = pd.DataFrame(...)  # Your sales data DataFrame
    >>> company_id = 1
    >>> quarter = 1
    >>> product_type = 'X'
    >>> result = aggregate_B2B_sales_by_grade(df, company_id, quarter, product_type)
    >>> print(result)
    ['(0) 65000.0']

    If there is only one item in the dictionary, result will be ['(0) 65000.0', 'Empty']:
    >>> df = pd.DataFrame(...)  # Your sales data DataFrame
    >>> company_id = 1
    >>> quarter = 1
    >>> product_type = 'X'
    >>> result = aggregate_B2B_sales_by_grade(df, company_id, quarter, product_type)
    >>> print(result)
    ['(0) 65000.0', 'Empty']
    
    If there are no items in the dictionary, result will be ['Empty']:
    >>> df = pd.DataFrame(...)  # Your sales data DataFrame
    >>> company_id = 1
    >>> quarter = 1
    >>> product_type = 'Z'  # Product that doesn't exist in the DataFrame
    >>> result = aggregate_B2B_sales_by_grade(df, company_id, quarter, product_type)
    >>> print(result)
    ['Empty']
    """
    # Filter the DataFrame based on the provided quarter, seller, and product
    registry = registry.get_quarter(quarter)
    filtered_df = registry[(registry['Seller'] == company.id) & (registry['Product'] == item)]

    # Group by 'Grade' and calculate the sum of volumes for each grade
    aggregated_data = filtered_df.groupby('Grade')['Volume'].sum().reset_index()

    # Convert the result to a dictionary
    result_dict = dict(zip(aggregated_data['Grade'], aggregated_data['Volume']))
    
    # Check the length of the dictionary and adjust the output accordingly
    if len(result_dict) == 1:
        grade, volume = result_dict.popitem()
        output = [f'({grade}) {volume}', 'Empty']
    elif len(result_dict) == 0:
        output = ['Empty', 'Empty']
    else:
        # Transform the dictionary into a list of strings
        output = [f'({key}) {value}' for key, value in result_dict.items()]

    return output



############################
# Applying borders helpers #
############################

def apply_border_to_column(ws, column, start_row, end_row, border):
    # Iterate through each row within the range
    for row in range(start_row, end_row + 1):
        cell = ws[f"{get_column_letter(column)}{row}"]
        cell.border = border

def apply_border_to_row(ws, row, start_column, end_column, border):
    # Iterate through each column within the range
    for col in range(start_column, end_column + 1):
        cell = ws[f"{get_column_letter(col)}{row}"]
        cell.border = cell.border + border

def apply_border_to_corner(cell, thickness, corner_type):
    border = copy(cell.border)

    # Define the side based on the border type
    if corner_type == "top-left":
        side = Side(border_style="thick")
        border.left = side
        border.top = side
    elif corner_type == "top-right":
        side = Side(border_style="thick")
        border.right = side
        border.top = side
    elif corner_type == "bottom-left":
        side = Side(border_style="thick")
        border.left = side
        border.bottom = side
    elif corner_type == "bottom-right":
        side = Side(border_style="thick")
        border.right = side
        border.bottom = side

    if thickness == "thick":
        thickness = 3
    elif thickness == "medium":
        thickness = 2
    else:
        thickness = 1

    cell.border = border

def apply_border_to_square(wb, start_column, start_row, num_columns, num_rows):
    ws = wb.active
    # Apply borders to side
    top_border = Border(top=Side(border_style="thick"))
    apply_border_to_row(ws, start_row, start_column, start_column + num_columns -1, top_border)
    bottom_border = Border(bottom=Side(border_style="thick"))
    apply_border_to_row(ws, start_row + num_rows - 1, start_column, start_column + num_columns -1, bottom_border)
    # then do column
    left_border = Border(left=Side(border_style="thick"))
    apply_border_to_column(ws, start_column, start_row, start_row + num_rows - 1, left_border)
    right_border = Border(right=Side(border_style="thick"))
    apply_border_to_column(ws, start_column + num_columns -1, start_row, start_row + num_rows - 1, right_border)
    
    # Finally, do the corners : 
    apply_border_to_corner(ws[f'{get_column_letter(start_column)}{start_row}'], "thick", corner_type='top-left')
    apply_border_to_corner(ws[f'{get_column_letter(start_column)}{start_row + num_rows -1}'], "thick", corner_type='bottom-left')
    apply_border_to_corner(ws[f'{get_column_letter(start_column + num_columns -1)}{start_row}'], "thick", corner_type='top-right')
    apply_border_to_corner(ws[f'{get_column_letter(start_column+ num_columns -1)}{start_row + num_rows -1}'], "thick", corner_type='bottom-right')

    return wb

def add_thin_border(cell:Cell or str):
    """
    Adds a thin border to all sides of the specified cell. \n
    Does not overwrite preexisting thick borders.

    Parameters
    ----------
    cell: Cell or str
        The openpyxl.cell.cell.Cell class or its str representation
    """
    old_border = copy(cell.border)
    new_border = Border(
        left=old_border.left if old_border.left and old_border.left.style == "thick" else Side(style="thin"),
        right=old_border.right if old_border.right and old_border.right.style == "thick" else Side(style="thin"),
        top=old_border.top if old_border.top and old_border.top.style == "thick" else Side(style="thin"),
        bottom=old_border.bottom if old_border.bottom and old_border.bottom.style == "thick" else Side(style="thin")
    )
    cell.border = new_border

def add_thin_borders_to_unit(wb, start_col, start_row):
    """
    Adds the required borders to each single unit
    """
    ws = wb.active
    #Inventory borders
    add_thin_border(ws[f'{get_column_letter(start_col)}{start_row}'])
    add_thin_border(ws[f'{get_column_letter(start_col+1)}{start_row}'])
    add_thin_border(ws[f'{get_column_letter(start_col)}{start_row+1}'])
    add_thin_border(ws[f'{get_column_letter(start_col+1)}{start_row+1}'])

    #Sales B2C borders
    add_thin_border(ws[f'{get_column_letter(start_col)}{start_row+3}'])
    add_thin_border(ws[f'{get_column_letter(start_col)}{start_row+4}'])
    add_thin_border(ws[f'{get_column_letter(start_col+1)}{start_row+3}'])
    add_thin_border(ws[f'{get_column_letter(start_col+1)}{start_row+4}'])

    #Sales B2B borders
    add_thin_border(ws[f'{get_column_letter(start_col)}{start_row+6}'])
    add_thin_border(ws[f'{get_column_letter(start_col)}{start_row+7}'])
    add_thin_border(ws[f'{get_column_letter(start_col+1)}{start_row+6}'])
    add_thin_border(ws[f'{get_column_letter(start_col+1)}{start_row+7}'])

    #Production borders
    add_thin_border(ws[f'{get_column_letter(start_col)}{start_row+9}'])
    add_thin_border(ws[f'{get_column_letter(start_col)}{start_row+10}'])
    add_thin_border(ws[f'{get_column_letter(start_col+1)}{start_row+9}'])
    add_thin_border(ws[f'{get_column_letter(start_col+1)}{start_row+10}'])

    #Max grades borders
    add_thin_border(ws[f'{get_column_letter(start_col)}{start_row+12}'])
    add_thin_border(ws[f'{get_column_letter(start_col)}{start_row+13}'])

    #Factories borders
    add_thin_border(ws[f'{get_column_letter(start_col)}{start_row+15}'])
    add_thin_border(ws[f'{get_column_letter(start_col)}{start_row+16}'])
    add_thin_border(ws[f'{get_column_letter(start_col+1)}{start_row+15}'])
    add_thin_border(ws[f'{get_column_letter(start_col+1)}{start_row+16}'])

    #Sales offices borders
    add_thin_border(ws[f'{get_column_letter(start_col)}{start_row+18}'])

if __name__ == "__main__":

    # Test the function with N = 3 and sample data as a 2D matrix
    MATRIX = [
        [1,None,None],
        [1,2,3],
        [1,2,3],
        [1,2,3],
        [1,2,3],
        [1,None,3],
        [1,2,3],
        [1,2,3],
        [1,2,3],
        [1,2,3]
    ]
    data = {
        'n_companies': 3,
        'n_regions': 2,
        'Quarter':3,
        'Company1' : {
            'Region1': MATRIX,
            'Region2': MATRIX,
            'Region3': MATRIX,
            'Region4': MATRIX
                      },
        'Company2' : {
            'Region1': MATRIX,
            'Region2': MATRIX,
            'Region3': MATRIX,
            'Region4': MATRIX
                      },
        'Company3' : {
            'Region1': MATRIX,
            'Region2': MATRIX,
            'Region3': MATRIX,
            'Region4': MATRIX
                      }
    }

    wb = generate_workbook(data)
    ws = wb.active
    add_thin_border(ws['B6'])
    add_thin_border(ws['C6'])
    add_thin_border(ws['D6'])
    

    # Cleaning up empty sheet
    try:
        wb.remove(wb['Sheet'])
    except:
        pass
    wb.save('output.xlsx')